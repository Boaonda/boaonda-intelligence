"""
Boaonda Intelligence — Processador de Metas Comerciais
=======================================================
Metas mensais de venda por representante (mercado interno), em PARES.

Fluxo (round-trip de Excel, mesmo padrão da Capacidade Fabril):

    Portal gera Excel pré-preenchido (meta global + sugestão ponderada)
      →  gestor edita offline (fixa reps, ajusta)
      →  importa  →  valida (formato + faixas)  →  backup do JSON  →  grava

A "sugestão" distribui a meta global do mês entre os representantes de forma
top-down, ponderando o histórico de cada um (regra do Gestor Comercial):
crescimento 1.5x, estável 1.0x, declínio 0.6x — sempre comparando com o
MESMO mês do ano anterior (deseasonaliza), nunca com o mês imediatamente
anterior. Reps marcados como "fixar" travam num número; o restante da meta
global se redistribui entre os não-fixados ("o resto se ajusta ao redor").

Fonte de realizado: dados_vendas_carteira.json (mesma base dos painéis de
Vendas — holdings × mês × espécie MI). Higiene idêntica ao frontend
(REPS_EXCLUIDOS / REP_UNIFICAR / apenas MI_PROG|PE|MISTA).

Saída: dados_metas.json
    {
      "gerado_em": "20/07/2026",
      "meta_global": { "202607": 150000 },
      "metas":       { "202607": { "CLEOBERTO ...": 12000, ... } },
      "detalhe":     { "202607": { "CLEOBERTO ...": {"tendencia": "...",
                                    "peso": 1.5, "fixado": true}, ... } }
    }
"""

import json, os, re, sys
from datetime import datetime
from collections import defaultdict

# ─── Higiene de representante (espelho de boaonda_carteira_representante.html) ─
REPS_EXCLUIDOS = {
    'ATEND. COMERCIAL - OUTRAS MARCAS',
    'ECOMMERCE',
}
REP_UNIFICAR = {
    'ALEXSHOES REPRESENTACOES LTDA':   'FRALDA & COMPANHIA LTDA',
    'ROGERIO RAPPA SILVEIRA - ME':     'RRAPPA REPRESENTACOES LTDA',
    'RIBEIRO UDI REPRESENTACOES LTDA': 'CLEOBERTO REPRESENTACOES LTDA',
}
# Só estas 3 espécies contam como MI válido (Programado/Pronta Entrega/Mista);
# exclui MI_ECOM (resíduo) e todos os canais ECOM_*/ME_*.
TIPOS_MI_VALIDOS = ['PROG', 'PE', 'MISTA']

# ─── Parâmetros de classificação (calibráveis) ────────────────────────────
PESO_TREND = {'crescimento': 1.5, 'estavel': 1.0, 'declinio': 0.6, 'novo': 1.0}
LIMIAR_CRESCIMENTO = 0.10   # > +10% vs mesmo período ano anterior
LIMIAR_DECLINIO    = -0.10  # < -10% vs mesmo período ano anterior
JANELA_REF_MESES   = 3      # nº de meses fechados usados como referência recente

_CORES = {  # MIV Boaonda
    'header_bg': '1C2030', 'header_ft': 'F3F0EB',
    'zebra_a': 'F5F5F5', 'zebra_b': 'F9F9F9', 'coral': 'ED6842',
    'travado_bg': 'FCE9E2',
}

# ─── Aritmética de competência (AAAAMM) ───────────────────────────────────
def mes_menos(aaaamm, n):
    """Subtrai n meses de uma competência 'AAAAMM'. Retorna 'AAAAMM'."""
    a, m = int(str(aaaamm)[:4]), int(str(aaaamm)[4:])
    total = a * 12 + (m - 1) - n
    return f'{total // 12:04d}{total % 12 + 1:02d}'


def mes_label(aaaamm):
    meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun',
             'jul', 'ago', 'set', 'out', 'nov', 'dez']
    s = str(aaaamm)
    return f'{meses[int(s[4:]) - 1]}/{s[:4]}'


# ─── Índice de realizado por representante ────────────────────────────────
def _carregar_carteira(carteira_path):
    with open(carteira_path, encoding='utf-8') as f:
        return json.load(f)


def indice_reps(carteira):
    """Constrói, a partir de dados_vendas_carteira.json:
        vol[rep][mes]    = pares MI válidos (soma das holdings do rep)
        ativos[rep][mes] = nº de holdings com pares > 0 no mês
        reps             = set de representantes normalizados (pós-higiene)
    """
    meta = carteira.get('meta', {})
    holdings = carteira.get('holdings', {})
    vol = defaultdict(lambda: defaultdict(float))
    ativos = defaultdict(lambda: defaultdict(int))
    reps = set()

    for h, hmeta in meta.items():
        r = (hmeta.get('representantes') or {}).get('MI')
        if not r or r in REPS_EXCLUIDOS:
            continue
        rep = REP_UNIFICAR.get(r, r)
        if rep in REPS_EXCLUIDOS:
            continue
        reps.add(rep)
        for mes, tipos in (holdings.get(h) or {}).items():
            tot = 0.0
            for t in TIPOS_MI_VALIDOS:
                par = tipos.get('MI_' + t)
                if par:
                    tot += par[0]
            if tot > 0:
                vol[rep][mes] += tot
                ativos[rep][mes] += 1
    return vol, ativos, reps


def _vol(vol, rep, meses):
    return sum(vol[rep].get(m, 0) for m in meses)


def _meses_fechados_antes(meses_disp, competencia, n=JANELA_REF_MESES):
    """Últimos n meses disponíveis estritamente anteriores à competência."""
    ant = sorted(m for m in meses_disp if m < competencia)
    return ant[-n:]


def classificar_rep(vol, rep, meses_disp, competencia):
    """Classifica a tendência do rep comparando os últimos meses fechados com
    o mesmo período do ano anterior. Retorna dict com tendencia/peso/base/…"""
    ref = _meses_fechados_antes(meses_disp, competencia)
    recente = _vol(vol, rep, ref)
    ref_ya = [mes_menos(m, 12) for m in ref]
    recente_ya = _vol(vol, rep, ref_ya)

    if recente_ya <= 0:
        tendencia = 'novo' if recente > 0 else 'novo'
        cresc = None
    else:
        cresc = (recente - recente_ya) / recente_ya
        if cresc > LIMIAR_CRESCIMENTO:
            tendencia = 'crescimento'
        elif cresc < LIMIAR_DECLINIO:
            tendencia = 'declinio'
        else:
            tendencia = 'estavel'

    # Base sazonalmente correta = mesmo mês do ano anterior; se 0, média
    # mensal dos últimos meses fechados (fallback para reps sem YoY do mês).
    base_yoy = _vol(vol, rep, [mes_menos(competencia, 12)])
    base = base_yoy if base_yoy > 0 else (recente / len(ref) if ref else 0)

    return {
        'tendencia': tendencia,
        'peso': PESO_TREND[tendencia],
        'realizado_3m': int(round(recente)),
        'base_yoy': int(round(base_yoy)),
        'base': base,
        'crescimento_yoy': cresc,
    }


def sugerir_distribuicao(carteira, competencia, meta_global, fixados=None):
    """Distribui meta_global (pares) entre os reps. `fixados` é um dict
    rep→pares travado; o restante (meta_global − Σ fixados) é rateado entre os
    não-fixados proporcionalmente a base×peso. Retorna (metas, detalhe)."""
    fixados = {k: v for k, v in (fixados or {}).items() if v is not None}
    vol, _ativos, reps = indice_reps(carteira)
    meses_disp = carteira.get('meses_disponiveis', [])

    info = {rep: classificar_rep(vol, rep, meses_disp, competencia) for rep in reps}

    flex = [r for r in reps if r not in fixados]
    soma_fix = sum(fixados.get(r, 0) for r in fixados)
    restante = max(0, meta_global - soma_fix)

    peso_total = sum(info[r]['base'] * info[r]['peso'] for r in flex)

    metas, detalhe = {}, {}
    for rep in sorted(reps):
        if rep in fixados:
            pares = int(round(fixados[rep]))
            fixado = True
        elif peso_total > 0:
            w = info[rep]['base'] * info[rep]['peso']
            pares = int(round(restante * w / peso_total))
            fixado = False
        else:
            pares = 0
            fixado = False
        metas[rep] = pares
        detalhe[rep] = {
            'tendencia': info[rep]['tendencia'],
            'peso': info[rep]['peso'],
            'realizado_3m': info[rep]['realizado_3m'],
            'base_yoy': info[rep]['base_yoy'],
            'fixado': fixado,
        }

    # Reconcilia o arredondamento no maior rep flexível para o total fechar
    # exatamente na meta global (promessa do LEIA-ME).
    if flex:
        soma_flex = sum(metas[r] for r in flex)
        diff = restante - soma_flex
        if diff:
            maior = max(flex, key=lambda r: metas[r])
            metas[maior] = max(0, metas[maior] + diff)
    return metas, detalhe


# ─── Estilos de célula (openpyxl) ─────────────────────────────────────────
def _cel(ws, row, col, valor, bold=False, bg=None, ft='000000',
         align='left', size=10, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, color=ft, name='Calibri', size=size)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c


def _header(ws, row, valores):
    for col, v in enumerate(valores, 1):
        _cel(ws, row, col, v, bold=True, bg=_CORES['header_bg'],
             ft=_CORES['header_ft'], align='center', wrap=True)


# ─── Abas do Excel ────────────────────────────────────────────────────────
_COLS_METAS = [
    'rep', 'realizado_3m', 'mesmo_mes_ano_passado', 'tendencia',
    'peso', 'clientes_ativos', 'sugestao_pares', 'fixar', 'meta_pares',
]
_COLS_OBRIG = ['rep', 'fixar', 'meta_pares']
_TEND_LABEL = {'crescimento': 'Crescimento', 'estavel': 'Estável',
               'declinio': 'Declínio', 'novo': 'Novo/sem histórico'}


def _sheet_leia_me(wb, competencia, meta_global):
    ws = wb.create_sheet('LEIA-ME')
    ws.column_dimensions['A'].width = 118
    linhas = [
        ('BOAONDA INTELLIGENCE — Metas Comerciais por Representante', True),
        (f'Competência: {mes_label(competencia)}   ·   Meta global: '
         f'{meta_global:,} pares'.replace(',', '.'), False),
        ('', False),
        ('COMO USAR', True),
        ('1. A aba PARÂMETROS traz a competência (mês) e a meta global de pares. '
         'Ajuste a meta global se quiser e reimporte para recalcular a sugestão.', False),
        ('2. Na aba METAS, cada linha é um representante. As colunas cinza são de '
         'apoio (só leitura) para você decidir com base em dado real.', False),
        ('3. Para TRAVAR a meta de um rep num valor específico: escreva "S" na '
         'coluna "fixar" e o número desejado em "meta_pares".', False),
        ('4. Os reps SEM "S" em "fixar" têm a meta recalculada automaticamente na '
         'importação — o restante da meta global (após os fixados) é redistribuído '
         'entre eles conforme o histórico (o resto se ajusta ao redor).', False),
        ('5. A soma das metas individuais PODE exceder a meta global (estratégia de '
         '"spread": atribuir aos reps mais do que a empresa precisa, criando folga '
         'na realização). Para isso, aumente a meta global (envelope) OU fixe reps '
         'com valores altos. Se os fixados já superam a meta global, os reps não '
         'fixados ficam em 0 — fixe também os que quiser com meta.', False),
        ('6. Salve e importe pelo portal (botão "Importar planilha" na tela de '
         'Metas Comerciais).', False),
        ('', False),
        ('ABA META EMPRESA — meta oficial da empresa por mês', True),
        ('  A meta da empresa é informada top-down e NÃO é a soma das metas dos '
         'representantes — pode ser diferente. Alimenta o quadro "Desempenho '
         'mensal" da operação.', False),
        ('  Preencha a coluna meta_empresa (pares) nos meses desejados — vem '
         'pré-preenchida com o último valor gravado, ou com a meta global do mês.', False),
        ('  Você pode informar vários meses de uma vez nessa aba.', False),
        ('', False),
        ('COLUNAS DA ABA METAS', True),
        ('  rep*                   — representante (não altere; é a chave)  OBRIGATÓRIO', False),
        ('  realizado_3m           — pares vendidos nos últimos 3 meses fechados (apoio)', False),
        ('  mesmo_mes_ano_passado  — pares no mesmo mês do ano anterior (base sazonal)', False),
        ('  tendencia              — Crescimento / Estável / Declínio (vs. ano anterior)', False),
        ('  peso                   — 1,5 crescimento · 1,0 estável · 0,6 declínio', False),
        ('  clientes_ativos        — nº de clientes com compra no último mês fechado', False),
        ('  sugestao_pares         — sugestão do sistema para a meta global informada', False),
        ('  fixar*                 — "S" trava a meta deste rep; vazio = recalculado  OBRIGATÓRIO', False),
        ('  meta_pares*            — meta em pares (usada quando fixar = "S")  OBRIGATÓRIO', False),
        ('', False),
        ('REGRAS DE VALIDAÇÃO', True),
        ('  meta global: inteiro > 0', False),
        ('  competência: AAAAMM (ex.: 202607)', False),
        ('  meta_pares: inteiro ≥ 0', False),
        ('  a soma das metas individuais pode exceder a meta global (spread liberado)', False),
        ('', False),
        ('METODOLOGIA (Gestor Comercial Boaonda)', True),
        ('  Base = mesmo mês do ano anterior (deseasonaliza). Tendência compara os '
         'últimos 3 meses fechados com o mesmo trimestre do ano anterior.', False),
        ('  Filtros aplicados: mercado interno, espécies Programado/Pronta Entrega/'
         'Mista; exclui e-commerce, EVA e exportação.', False),
    ]
    for i, (txt, bold) in enumerate(linhas, 1):
        _cel(ws, i, 1, txt, bold=bold,
             bg=_CORES['header_bg'] if bold else None,
             ft=_CORES['header_ft'] if bold else '000000',
             size=11 if bold else 10, wrap=True)


def _sheet_parametros(wb, competencia, meta_global):
    ws = wb.create_sheet('PARÂMETROS')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 46
    _header(ws, 1, ['CHAVE', 'VALOR', 'DESCRIÇÃO'])
    linhas = [
        ('competencia', int(competencia), 'Mês da meta no formato AAAAMM (ex.: 202607)'),
        ('meta_global', int(meta_global), 'Meta total de pares do mês (será distribuída)'),
    ]
    for i, (k, v, d) in enumerate(linhas, 2):
        _cel(ws, i, 1, k)
        _cel(ws, i, 2, v, align='center')
        _cel(ws, i, 3, d)
    ws.freeze_panes = 'A2'


def _sheet_metas(wb, carteira, competencia, meta_global, metas_atuais=None):
    ws = wb.create_sheet('METAS')
    larguras = [34, 14, 20, 14, 8, 14, 14, 8, 12]
    for col, w in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
    _header(ws, 1, _COLS_METAS)

    vol, ativos, reps = indice_reps(carteira)
    meses_disp = carteira.get('meses_disponiveis', [])
    ref_yoy = mes_menos(competencia, 12)
    ult_fechado = (_meses_fechados_antes(meses_disp, competencia) or [None])[-1]

    # Sugestão pura (sem fixados) para a meta global informada.
    metas_atuais = metas_atuais or {}
    fixados_ini = {r: metas_atuais[r] for r in metas_atuais} if metas_atuais else None
    sugestao, _det = sugerir_distribuicao(carteira, competencia, meta_global)

    for i, rep in enumerate(sorted(reps), 2):
        info = classificar_rep(vol, rep, meses_disp, competencia)
        cli_at = ativos[rep].get(ult_fechado, 0) if ult_fechado else 0
        yoy_pares = int(round(_vol(vol, rep, [ref_yoy])))
        bg = _CORES['zebra_a'] if i % 2 == 0 else None
        fixado = rep in metas_atuais
        meta_pre = metas_atuais.get(rep, sugestao.get(rep, 0))
        row = [
            (rep, 'left', False),
            (info['realizado_3m'], 'center', True),
            (f'{mes_label(ref_yoy)}: {yoy_pares:,}'.replace(',', '.'), 'center', True),
            (_TEND_LABEL[info['tendencia']], 'center', True),
            (info['peso'], 'center', True),
            (cli_at, 'center', True),
            (sugestao.get(rep, 0), 'center', True),
            ('S' if fixado else '', 'center', False),
            (int(round(meta_pre)), 'center', False),
        ]
        for col, (val, align, apoio) in enumerate(row, 1):
            cell_bg = _CORES['zebra_b'] if apoio else (
                _CORES['travado_bg'] if (col >= 8 and fixado) else bg)
            _cel(ws, i, col, val, bg=cell_bg, align=align,
                 ft='6b6b6b' if apoio else '000000')
    ws.freeze_panes = 'B2'


def _meses_meta_empresa(carteira):
    """Janela de meses oferecida na aba META EMPRESA: últimos 13 meses de
    dados + próximos 3 (para planejar meses ainda sem venda)."""
    meses = sorted(carteira.get('meses_disponiveis', []))
    if not meses:
        base = datetime.now().strftime('%Y%m')
        return [mes_menos(base, -n) for n in range(3, -4, -1)]
    ult = meses[-1]
    janela = [mes_menos(ult, n) for n in range(12, -4, -1)]  # ult-12 .. ult+3
    return janela


def _sheet_meta_empresa(wb, carteira, meta_empresa_atual, meta_global_atual):
    """Aba META EMPRESA — meta oficial da empresa por mês (um valor por linha).
    É informada top-down e NÃO é a soma das metas dos representantes; alimenta
    o quadro 'Desempenho mensal' da operação. Pré-preenche com o valor já
    gravado ou, na falta, com o meta_global (envelope distribuído) do mês."""
    ws = wb.create_sheet('META EMPRESA')
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40
    _header(ws, 1, ['mes', 'meta_empresa', 'referência'])
    for i, m in enumerate(_meses_meta_empresa(carteira), 2):
        val = meta_empresa_atual.get(m)
        if val in (None, ''):
            val = meta_global_atual.get(m, '')
        bg = _CORES['zebra_a'] if i % 2 == 0 else None
        _cel(ws, i, 1, int(m), bg=bg, align='center')
        _cel(ws, i, 2, int(val) if val not in (None, '') else '', bg=bg, align='center')
        _cel(ws, i, 3, mes_label(m), bg=bg, ft='6b6b6b')
    ws.freeze_panes = 'A2'


# ─── Exportar ─────────────────────────────────────────────────────────────
def exportar_metas_excel(dados_metas_path, carteira_path, competencia, meta_global):
    """Gera o Excel de metas (BytesIO). Se já houver metas gravadas para a
    competência, elas voltam pré-preenchidas (com fixar = "S")."""
    from io import BytesIO
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError('openpyxl não instalado. Rode: pip install openpyxl')

    competencia = str(competencia)
    meta_global = int(meta_global)
    carteira = _carregar_carteira(carteira_path)

    metas_atuais, meta_global_all, meta_empresa_all = {}, {}, {}
    if os.path.exists(dados_metas_path):
        with open(dados_metas_path, encoding='utf-8') as f:
            d = json.load(f)
        metas_atuais = (d.get('metas') or {}).get(competencia, {}) or {}
        meta_global_all = d.get('meta_global') or {}
        meta_empresa_all = d.get('meta_empresa') or {}
        if not meta_global:
            meta_global = int(meta_global_all.get(competencia, 0))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_leia_me(wb, competencia, meta_global)
    _sheet_parametros(wb, competencia, meta_global)
    _sheet_metas(wb, carteira, competencia, meta_global, metas_atuais)
    _sheet_meta_empresa(wb, carteira, meta_empresa_all, meta_global_all)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Importar ─────────────────────────────────────────────────────────────
def _norm_header(v):
    s = str(v or '').strip().lower()
    return s.split('\n')[0].rstrip('*').strip()


def _achar_header_metas(ws):
    linhas = list(ws.iter_rows(max_row=10))
    for r_idx, row in enumerate(linhas, 1):
        header = [_norm_header(c.value) for c in row]
        if 'rep' in header and 'meta_pares' in header:
            return r_idx, header
    return 1, ([_norm_header(c.value) for c in linhas[0]] if linhas else [])


def _ler_parametros(ws):
    params = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and str(row[0]).strip():
            params[str(row[0]).strip()] = row[1]
    return params


def _validar(wb):
    """Camadas 1 (formato) e 2 (faixas). Retorna lista de erros."""
    erros = []
    for sheet in ('PARÂMETROS', 'METAS'):
        if sheet not in wb.sheetnames:
            erros.append(f'Aba "{sheet}" não encontrada.')
    if erros:
        return erros

    params = _ler_parametros(wb['PARÂMETROS'])
    comp = str(params.get('competencia', '')).strip()
    if not (comp.isdigit() and len(comp) == 6):
        erros.append(f'PARÂMETROS › competencia inválida: "{comp}" (esperado AAAAMM)')
    try:
        mg = int(float(params.get('meta_global')))
        if mg <= 0:
            erros.append(f'PARÂMETROS › meta_global deve ser > 0 (recebido {mg})')
    except (TypeError, ValueError):
        erros.append(f'PARÂMETROS › meta_global inválida: "{params.get("meta_global")}"')
        mg = 0

    ws = wb['METAS']
    header_row, header = _achar_header_metas(ws)
    for col in _COLS_OBRIG:
        if col not in header:
            erros.append(f'Coluna obrigatória ausente em METAS: {col}')
    if erros:
        return erros

    # As metas individuais PODEM exceder a meta global (estratégia de spread:
    # atribuir aos reps mais do que a empresa precisa, criando folga na
    # realização). Portanto não há trava sobre a soma dos fixados — só se
    # valida que cada meta_pares é um número não negativo.
    idx = {n: i for i, n in enumerate(header)}
    for n_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                header_row + 1):
        if not row or row[idx['rep']] is None or not str(row[idx['rep']]).strip():
            continue
        rep = str(row[idx['rep']]).strip()
        raw = row[idx['meta_pares']]
        if raw in (None, ''):
            continue
        try:
            mp = float(raw)
        except (TypeError, ValueError):
            erros.append(f'METAS › linha {n_row} ({rep}) › meta_pares não numérico: "{raw}"')
            continue
        if mp < 0:
            erros.append(f'METAS › linha {n_row} ({rep}) › meta_pares negativo: {mp}')
    return erros


def importar_metas_excel(path_excel, dados_metas_path, carteira_path, output_dir=None):
    """Importa o Excel, recalcula a distribuição (fixados + redistribuição) e
    grava dados_metas.json (com backup). Retorna dict status/detalhes."""
    try:
        import openpyxl
    except ImportError:
        return {'status': 'erro', 'mensagem': 'openpyxl não instalado.'}

    if output_dir is None:
        output_dir = os.path.dirname(dados_metas_path)

    wb = openpyxl.load_workbook(path_excel, data_only=True)

    erros = _validar(wb)
    if erros:
        return {'status': 'erro', 'mensagem': 'Falha na validação:\n' + '\n'.join(erros)}

    params = _ler_parametros(wb['PARÂMETROS'])
    competencia = str(params['competencia']).strip()
    meta_global = int(float(params['meta_global']))

    ws = wb['METAS']
    header_row, header = _achar_header_metas(ws)
    idx = {n: i for i, n in enumerate(header)}
    fixados = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[idx['rep']] is None or not str(row[idx['rep']]).strip():
            continue
        rep = str(row[idx['rep']]).strip()
        fixar = str(row[idx['fixar']] or '').strip().upper()
        if fixar in ('S', 'SIM', '1', 'TRUE', 'VERDADEIRO', 'X'):
            raw = row[idx['meta_pares']]
            try:
                fixados[rep] = int(round(float(raw))) if raw not in (None, '') else 0
            except (TypeError, ValueError):
                fixados[rep] = 0

    carteira = _carregar_carteira(carteira_path)
    metas, detalhe = sugerir_distribuicao(carteira, competencia, meta_global, fixados)

    # Aba META EMPRESA (opcional) — meta oficial da empresa por mês
    meta_empresa_import = {}
    if 'META EMPRESA' in wb.sheetnames:
        wsE = wb['META EMPRESA']
        for row in wsE.iter_rows(min_row=2, values_only=True):
            if not row or row[0] in (None, ''):
                continue
            mes = str(row[0]).strip()
            if not (mes.isdigit() and len(mes) == 6):
                continue
            val = row[1] if len(row) > 1 else None
            if val in (None, ''):
                continue
            try:
                meta_empresa_import[mes] = int(round(float(val)))
            except (TypeError, ValueError):
                pass

    # Backup + merge da competência no JSON existente
    dados = {'meta_global': {}, 'metas': {}, 'detalhe': {}, 'meta_empresa': {}}
    if os.path.exists(dados_metas_path):
        with open(dados_metas_path, encoding='utf-8') as f:
            dados = json.load(f)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        with open(os.path.join(output_dir, f'dados_metas_backup_{ts}.json'),
                  'w', encoding='utf-8') as f:
            json.dump(dados, f, ensure_ascii=False, default=str)

    dados.setdefault('meta_global', {})[competencia] = meta_global
    dados.setdefault('metas', {})[competencia] = metas
    dados.setdefault('detalhe', {})[competencia] = detalhe
    # meta_empresa: se a aba não trouxe o mês, usa o meta_global como padrão
    dados.setdefault('meta_empresa', {}).update(meta_empresa_import)
    dados['meta_empresa'].setdefault(competencia, meta_global)
    dados['gerado_em'] = datetime.now().strftime('%d/%m/%Y')

    with open(dados_metas_path, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, default=str)

    n_fix = sum(1 for r in detalhe.values() if r['fixado'])
    return {
        'status': 'ok',
        'competencia': competencia,
        'competencia_label': mes_label(competencia),
        'meta_global': meta_global,
        'reps': len(metas),
        'reps_fixados': n_fix,
        'total_distribuido': sum(metas.values()),
        'meta_empresa': dados['meta_empresa'].get(competencia),
        'meses_meta_empresa': len(meta_empresa_import) or None,
        'gerado_em': dados['gerado_em'],
    }


# ═══ REGISTRO FORMAL EM GRADE (planilha das Configurações, admin) ═════════
# Uma planilha única: linha META EMPRESA + uma linha por representante,
# colunas = todos os meses com vendas. A planilha é a FONTE COMPLETA: ao
# importar, dados_metas.json é reconstruído a partir dela (célula em branco =
# sem meta). Sem sugestão/ponderação aqui — isso é o "apoio para formar metas"
# (rascunho), separado do registro formal.
_MESES_ABREV = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun',
                'jul', 'ago', 'set', 'out', 'nov', 'dez']
_LABEL_META_EMPRESA = 'META EMPRESA'


def _parse_mes_header(v):
    """Reconhece cabeçalho de mês: 'jan/2025', 'jan/25' ou '202501'."""
    s = str(v or '').strip().lower()
    if s.isdigit() and len(s) == 6:
        return s
    m = re.match(r'([a-zç]{3})[./\- ]?(\d{2,4})$', s)
    if m and m.group(1) in _MESES_ABREV:
        mo = _MESES_ABREV.index(m.group(1)) + 1
        yr = int(m.group(2))
        yr = 2000 + yr if yr < 100 else yr
        return f'{yr:04d}{mo:02d}'
    return None


def _sheet_leia_me_grade(wb, meses):
    ws = wb.create_sheet('LEIA-ME')
    ws.column_dimensions['A'].width = 118
    ini = mes_label(meses[0]) if meses else '-'
    fim = mes_label(meses[-1]) if meses else '-'
    linhas = [
        ('BOAONDA INTELLIGENCE — Registro de Metas Comerciais', True),
        (f'Planilha única com todos os representantes e a meta da empresa, por mês '
         f'({ini} a {fim}).', False),
        ('', False),
        ('COMO USAR', True),
        ('1. Na aba METAS, cada coluna é um mês e cada linha é um representante.', False),
        ('2. A linha "META EMPRESA" (no topo) é a meta global da empresa por mês — '
         'informada top-down; NÃO é a soma das metas dos representantes.', False),
        ('3. Cada célula é a meta em pares daquele rep/mês. A soma das metas dos reps '
         'pode ser maior ou menor que a meta da empresa (spread liberado).', False),
        ('4. Célula em branco = SEM meta para aquele rep/mês (removida na importação). '
         'A planilha é a FONTE COMPLETA: o que estiver nela é o que fica gravado.', False),
        ('5. Não altere a linha de cabeçalho (os meses) nem os nomes dos '
         'representantes na coluna A.', False),
        ('6. Salve e importe pelo portal: Configurações › Metas comerciais › '
         'Importar planilha.', False),
        ('', False),
        ('REGRAS', True),
        ('  valores: inteiros ≥ 0 (pares); vazio = sem meta', False),
        ('  cabeçalho dos meses no formato "jan/2025" — não editar', False),
    ]
    for i, (t, b) in enumerate(linhas, 1):
        _cel(ws, i, 1, t, bold=b, bg=_CORES['header_bg'] if b else None,
             ft=_CORES['header_ft'] if b else '000000', size=11 if b else 10, wrap=True)


def exportar_metas_grade(dados_metas_path, carteira_path):
    """Planilha em grade (reps × meses + linha META EMPRESA), pré-preenchida
    com o que está gravado. Retorna BytesIO."""
    from io import BytesIO
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError('openpyxl não instalado. Rode: pip install openpyxl')

    carteira = _carregar_carteira(carteira_path)
    meses = sorted(carteira.get('meses_disponiveis', []))
    _v, _a, reps = indice_reps(carteira)
    reps = sorted(reps)

    dados = {}
    if os.path.exists(dados_metas_path):
        with open(dados_metas_path, encoding='utf-8') as f:
            dados = json.load(f)
    meta_empresa = dados.get('meta_empresa', {}) or {}
    metas = dados.get('metas', {}) or {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_leia_me_grade(wb, meses)
    ws = wb.create_sheet('METAS')
    ws.column_dimensions['A'].width = 34
    _cel(ws, 1, 1, 'Representante', bold=True, bg=_CORES['header_bg'],
         ft=_CORES['header_ft'])
    for j, m in enumerate(meses, 2):
        ws.column_dimensions[ws.cell(1, j).column_letter].width = 11
        _cel(ws, 1, j, mes_label(m), bold=True, bg=_CORES['header_bg'],
             ft=_CORES['header_ft'], align='center')

    _cel(ws, 2, 1, _LABEL_META_EMPRESA, bold=True, bg=_CORES['travado_bg'])
    for j, m in enumerate(meses, 2):
        v = meta_empresa.get(m)
        _cel(ws, 2, j, int(v) if v not in (None, '') else '',
             bg=_CORES['travado_bg'], align='center')

    for i, rep in enumerate(reps, 3):
        bg = _CORES['zebra_a'] if i % 2 == 0 else None
        _cel(ws, i, 1, rep, bg=bg)
        for j, m in enumerate(meses, 2):
            v = (metas.get(m) or {}).get(rep)
            _cel(ws, i, j, int(v) if v not in (None, '') else '', bg=bg, align='center')
    ws.freeze_panes = 'B3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def importar_metas_grade(path_excel, dados_metas_path, output_dir=None):
    """Reconstrói dados_metas.json a partir da grade (fonte completa; branco =
    sem meta). Faz backup do anterior. Retorna dict status/detalhes."""
    try:
        import openpyxl
    except ImportError:
        return {'status': 'erro', 'mensagem': 'openpyxl não instalado.'}
    if output_dir is None:
        output_dir = os.path.dirname(dados_metas_path)

    wb = openpyxl.load_workbook(path_excel, data_only=True)
    nome_aba = 'METAS' if 'METAS' in wb.sheetnames else wb.sheetnames[-1]
    ws = wb[nome_aba]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return {'status': 'erro', 'mensagem': 'Planilha vazia.'}

    col_mes = {}
    for ci, hv in enumerate(linhas[0]):
        if ci == 0:
            continue
        mm = _parse_mes_header(hv)
        if mm:
            col_mes[ci] = mm
    if not col_mes:
        return {'status': 'erro',
                'mensagem': 'Nenhuma coluna de mês reconhecida no cabeçalho '
                            '(esperado "jan/2025", "202501", …).'}

    meta_empresa, metas, erros = {}, {}, []
    for ri, row in enumerate(linhas[1:], 2):
        if not row or row[0] is None or not str(row[0]).strip():
            continue
        nome = str(row[0]).strip()
        eh_empresa = nome.upper().replace(' ', '') == _LABEL_META_EMPRESA.replace(' ', '')
        for ci, mm in col_mes.items():
            raw = row[ci] if ci < len(row) else None
            if raw in (None, ''):
                continue
            try:
                v = int(round(float(raw)))
            except (TypeError, ValueError):
                erros.append(f'linha {ri} ({nome}) › {mes_label(mm)}: valor não numérico "{raw}"')
                continue
            if v < 0:
                erros.append(f'linha {ri} ({nome}) › {mes_label(mm)}: valor negativo {v}')
                continue
            if eh_empresa:
                meta_empresa[mm] = v
            else:
                metas.setdefault(mm, {})[nome] = v
    if erros:
        return {'status': 'erro', 'mensagem': 'Falha na validação:\n' + '\n'.join(erros[:30])}

    metas = {m: d for m, d in metas.items() if d}  # descarta meses sem nenhum rep

    if os.path.exists(dados_metas_path):
        with open(dados_metas_path, encoding='utf-8') as f:
            antigo = json.load(f)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        with open(os.path.join(output_dir, f'dados_metas_backup_{ts}.json'),
                  'w', encoding='utf-8') as f:
            json.dump(antigo, f, ensure_ascii=False, default=str)

    dados = {
        'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'meta_empresa': meta_empresa,
        'metas': metas,
    }
    with open(dados_metas_path, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, default=str)

    reps_com_meta = len({r for d in metas.values() for r in d})
    return {
        'status': 'ok',
        'meses': len(col_mes),
        'meses_meta_empresa': len(meta_empresa),
        'reps_com_meta': reps_com_meta,
        'celulas': sum(len(d) for d in metas.values()),
        'gerado_em': dados['gerado_em'],
    }


# ─── CLI (teste rápido) ───────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print('uso: python processador_metas.py <competencia AAAAMM> <meta_global> '
              '[dir=frontend]')
        sys.exit(1)
    competencia, meta_global = sys.argv[1], int(sys.argv[2])
    d = sys.argv[3] if len(sys.argv) > 3 else 'frontend'
    carteira_path = os.path.join(d, 'dados_vendas_carteira.json')
    carteira = _carregar_carteira(carteira_path)
    metas, detalhe = sugerir_distribuicao(carteira, competencia, meta_global)
    print(f'\n  Sugestão de metas — {mes_label(competencia)} — '
          f'meta global {meta_global:,} pares\n'.replace(',', '.'))
    print(f'  {"REPRESENTANTE":<36}{"TEND":<14}{"PESO":>5}{"META":>10}')
    print('  ' + '-' * 66)
    for rep in sorted(metas, key=lambda r: -metas[r]):
        dd = detalhe[rep]
        print(f'  {rep[:34]:<36}{_TEND_LABEL[dd["tendencia"]][:12]:<14}'
              f'{dd["peso"]:>5}{metas[rep]:>10,}'.replace(',', '.'))
    print('  ' + '-' * 66)
    print(f'  {"TOTAL":<55}{sum(metas.values()):>10,}'.replace(',', '.'))


if __name__ == '__main__':
    main()
