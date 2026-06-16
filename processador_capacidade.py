"""
Boaonda Intelligence — Processador de Capacidade Fabril
========================================================
Lê a planilha "Programação_GERAL" (aba CAP_PROD DIA) e gera
dados_capacidade.json com a capacidade produtiva por referência/linha
(Camada 1), o pool de máquinas de injetados (Camada 2), o teto agregado
do atelier (Camada 0) e os setores de apoio (Palmilha, Sola, Pintura).

Bloco independente — sem cruzamento com vendas/estoque/programação
nesta fase, exceto a validação de cobertura (seção 4 da especificação),
que cruza as referências de `dados_refs_tabela.json` apenas para
medir cobertura, sem alterar os dados de capacidade.

Ver: ESPECIFICACAO_motor_capacidade.md

Entrada:
    - Programação_GERAL.xlsx, aba "CAP_PROD DIA", linhas 9+,
      colunas B-L (Camada 1) e AA/AB/AD/AF-AL (Camada 2)

Saída (gravada em output_dir):
    - dados_capacidade.json
"""

import json, os, sys
from datetime import datetime
from collections import Counter, defaultdict

# ─── CAMADA 0 — Teto do Atelier (semanal, agregado) ────────────────────
TETO_ATELIER_SEMANAL = {
    'convencional_pares_semana': 29000,
    'montado_pares_semana': 6000,
    'total_pares_semana': 35000,
    'observacao': (
        'Limite agregado da soma de TODAS as referências do mix programado '
        'na semana, independente das capacidades individuais por referência '
        '(Camada 1).'
    ),
}

# ─── CAMADA 2 — Pool de Máquinas de Injetados (config do parque) ───────
POOL_MAQUINAS_INJETADOS = {
    'horizontal': {'maquinas': 9, 'horas_dia': 8.48, 'min_dia_total': 4579.2},
    'tiras':      {'maquinas': 3, 'horas_dia': 8.48, 'min_dia_total': 1526.4},
    'rotativa':   {'maquinas': 1, 'horas_dia': 23,   'min_dia_total': 1380.0},
    'vertical':   {'maquinas': 1, 'horas_dia': 8.48, 'min_dia_total': 508.8},
    'eva': {
        'maquinas': 3, 'horas_dia': 24, 'min_dia_total': 4320.0,
        'maquinas_ativas': ['maq1', 'maq2', 'maq3'],
        'maquina_reserva': 'maq4 (não incluída na capacidade padrão)',
    },
}

EFICIENCIAS_PADRAO = {
    'horizontal': 0.60, 'tiras': 0.60, 'vertical': 0.55,
    'rotativa_outros': 0.50, 'rotativa_lily': 0.55, 'rotativa_nellie': 0.70,
    'eva_maq1': 0.80, 'eva_maq2': 0.75, 'eva_maq3': 0.75, 'eva_maq4': 1.0,
    'eva_maq4_observacao': (
        'Máquina reserva, NÃO incluída no cálculo padrão do pool EVA'
    ),
}

GIRO_DIAS = {
    'corte': 7, 'costura': 4, 'palmilha': 3,
    'observacao': 'Lead time em dias úteis antes da montagem (sequencial, não simultâneo)',
}

# ─── Mapeamento de colunas da aba CAP_PROD DIA (1-indexado, A=1) ───────
COL = {
    'ref_linha': 2, 'material': 3, 'referencia': 4, 'linha': 5,
    'descricao': 6, 'tipo_montagem': 7,
    'cap_mont1_conv': 8, 'cap_mont2_montado': 9,
    'cap_palmilha': 10, 'cap_sola': 11, 'cap_pintura': 12,
    # Camada 2 — tempos padrão do pool de injetados (min/par)
    'horiz_sola': 27, 'horiz_tiras': 28, 'vert_sola': 30,
    'rot_outros': 32, 'rot_lily': 33, 'rot_nellie': 34,
    'eva_maq1': 35, 'eva_maq2': 36, 'eva_maq3': 37, 'eva_maq4': 38,
}

# Coluna planilha -> grupo do pool de injetados
COLUNA_POOL = {
    'horiz_sola': 'horizontal', 'horiz_tiras': 'tiras', 'vert_sola': 'vertical',
    'rot_outros': 'rotativa', 'rot_lily': 'rotativa', 'rot_nellie': 'rotativa',
    'eva_maq1': 'eva', 'eva_maq2': 'eva', 'eva_maq3': 'eva', 'eva_maq4': 'eva',
}

LINHA_INICIO = 9
ABA_CAPACIDADE = 'CAP_PROD DIA'

# Capacidade observada (PRS/DIA) dos setores — aba ANALISE da planilha,
# levantamento de 12/06/2026. Apenas informativa (Bloco 1 do dashboard);
# atualizar manualmente se o usuário remedir esses valores.
CAPACIDADE_DIA_OBSERVADA = {
    'mont1_convencional': 7302.18,
    'mont2_montado': 1085.48,
    'filial_palmilha': 4394.84,
    'filial_sola': 3313.67,
    'pintura': 1197.31,
}


def _num(v):
    """Converte célula da planilha em float, tratando vazio/None."""
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def processar_capacidade(caminho_planilha, output_dir='.'):
    print("\n  Processando capacidade fabril...")
    try:
        import openpyxl
    except ImportError:
        print("    ✗ openpyxl não instalado. Rode: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(caminho_planilha, data_only=True, read_only=True)
    ws = wb[ABA_CAPACIDADE]

    referencias = {}
    por_tipo = Counter()
    palmilha_caps, sola_caps = set(), set()
    palmilha_refs = sola_refs = 0
    pintura_grupos = Counter()

    for row in ws.iter_rows(min_row=LINHA_INICIO, values_only=True):
        ref_linha = row[COL['ref_linha'] - 1]
        if ref_linha is None:
            continue
        ref_linha = str(ref_linha).strip()
        if not ref_linha:
            continue

        tipo_raw = str(row[COL['tipo_montagem'] - 1] or '').strip().upper()
        if 'CONVENCIONAL' in tipo_raw:
            tipo = 'CONVENCIONAL'
        elif 'MONTADO' in tipo_raw:
            tipo = 'MONTADO'
        else:
            tipo = None

        cap_conv = _num(row[COL['cap_mont1_conv'] - 1])
        cap_mont = _num(row[COL['cap_mont2_montado'] - 1])
        cap_montagem = cap_conv if tipo == 'CONVENCIONAL' else (cap_mont if tipo == 'MONTADO' else None)
        if not cap_montagem:
            continue  # só referências com capacidade de montagem definida (Camada 1)

        por_tipo[tipo] += 1

        cap_palmilha = _num(row[COL['cap_palmilha'] - 1])
        cap_sola = _num(row[COL['cap_sola'] - 1])
        cap_pintura = _num(row[COL['cap_pintura'] - 1])
        if cap_palmilha:
            palmilha_refs += 1
            palmilha_caps.add(int(cap_palmilha))
        if cap_sola:
            sola_refs += 1
            sola_caps.add(int(cap_sola))
        if cap_pintura:
            pintura_grupos[int(cap_pintura)] += 1

        tempos = defaultdict(list)
        for col_nome, pool in COLUNA_POOL.items():
            v = _num(row[COL[col_nome] - 1])
            if v:
                tempos[pool].append({'componente': col_nome, 'tempo_padrao_min': round(v, 2)})

        referencias[ref_linha] = {
            'material': str(row[COL['material'] - 1] or '').strip(),
            'referencia': str(row[COL['referencia'] - 1] or '').strip(),
            'linha': row[COL['linha'] - 1],
            'descricao': str(row[COL['descricao'] - 1] or '').strip(),
            'tipo_montagem': tipo,
            'capacidade_montagem_dia': int(cap_montagem),
            'capacidades_apoio': {
                'palmilha': int(cap_palmilha) if cap_palmilha else None,
                'sola': int(cap_sola) if cap_sola else None,
                'pintura': int(cap_pintura) if cap_pintura else None,
            },
            'tempos_padrao_pool_min': dict(tempos),
        }

    print(f"    {len(referencias)} referências com capacidade de montagem "
          f"({por_tipo.get('CONVENCIONAL', 0)} conv. + {por_tipo.get('MONTADO', 0)} montado)")

    setores_apoio = {
        'filial_palmilha': {
            'capacidade_dia': max(palmilha_caps) if palmilha_caps else 0,
            'refs_que_usam': palmilha_refs,
            'observacao': 'Capacidade do setor inteiro, idêntica em todas as refs que passam por ele',
        },
        'filial_sola': {
            'capacidade_dia': max(sola_caps) if sola_caps else 0,
            'refs_que_usam': sola_refs,
            'observacao': 'Capacidade do setor inteiro, idêntica em todas as refs que passam por ele',
        },
        'pintura': {
            'grupos': {str(k): v for k, v in sorted(pintura_grupos.items())},
            'total_refs': sum(pintura_grupos.values()),
            'observacao': 'Capacidade por linha/célula de pintura, transversal a CONVENCIONAL e MONTADO',
        },
        'giro_dias': GIRO_DIAS,
        'capacidade_dia_observada': CAPACIDADE_DIA_OBSERVADA,
    }

    metadados = {
        'total_referencias_cadastradas': sum(1 for row in ws.iter_rows(min_row=LINHA_INICIO, values_only=True)
                                              if row[COL['ref_linha'] - 1]),
        'total_com_capacidade_montagem': len(referencias),
        'por_tipo': dict(por_tipo),
    }

    validacao_cobertura = calcular_validacao_cobertura(referencias, output_dir)

    dados = {
        'gerado_em': datetime.now().strftime('%d/%m/%Y'),
        'teto_atelier_semanal': TETO_ATELIER_SEMANAL,
        'pool_maquinas_injetados': POOL_MAQUINAS_INJETADOS,
        'eficiencias_padrao': EFICIENCIAS_PADRAO,
        'metadados': metadados,
        'referencias': referencias,
        'setores_apoio': setores_apoio,
        'validacao_cobertura': validacao_cobertura,
    }

    with open(os.path.join(output_dir, 'dados_capacidade.json'), 'w', encoding='utf-8') as f_:
        json.dump(dados, f_, ensure_ascii=False, default=str)
    print(f"    ✓ dados_capacidade.json gerado ({len(referencias)} referências)")
    return dados


def calcular_validacao_cobertura(referencias, output_dir):
    """Cruza as referências com capacidade cadastrada contra o volume real
    da programação (dados_refs_tabela.json) — única referência cruzada
    permitida nesta fase (seção 4 da especificação)."""
    refs_capacidade = {v['referencia'] for v in referencias.values()}

    pares_por_ref = Counter()
    try:
        with open(os.path.join(output_dir, 'dados_refs_tabela.json'), encoding='utf-8') as f_:
            rt = json.load(f_)
        for semana in (rt.get('semanas') or {}).values():
            for ref, qtd in (semana.get('refs') or {}).items():
                pares_por_ref[ref] += qtd
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    refs_programacao = set(pares_por_ref.keys())
    cobertas = refs_programacao & refs_capacidade
    nao_cobertas = refs_programacao - refs_capacidade
    total_pares = sum(pares_por_ref.values())
    pares_cobertos = sum(p for r, p in pares_por_ref.items() if r in refs_capacidade)

    return {
        'refs_programacao_unicas': len(refs_programacao),
        'refs_com_capacidade': len(cobertas),
        'total_pares_periodo': total_pares,
        'pares_cobertos': pares_cobertos,
        'pct_cobertura_volume': round(pares_cobertos / total_pares * 100, 1) if total_pares else None,
        'refs_sem_capacidade': sorted(
            ({'referencia': r, 'pares': pares_por_ref[r]} for r in nao_cobertas),
            key=lambda x: -x['pares']
        ),
    }


def achar_planilha_capacidade(diretorio='.'):
    """Procura a planilha 'Programação_GERAL' (.xlsx) na pasta indicada."""
    for nome in os.listdir(diretorio):
        nome_lower = nome.lower()
        if nome_lower.endswith('.xlsx') and 'program' in nome_lower and 'geral' in nome_lower:
            return os.path.join(diretorio, nome)
    return None


# ─── Exportar / Importar Excel ───────────────────────────────────────

_POOLS_ORDEM = ['horizontal', 'tiras', 'rotativa', 'vertical', 'eva']
_EFIC_ORDEM  = ['horizontal', 'tiras', 'vertical',
                 'rotativa_outros', 'rotativa_lily', 'rotativa_nellie',
                 'eva_maq1', 'eva_maq2', 'eva_maq3']
_COLS_REFS   = [
    'ref_linha', 'referencia', 'linha', 'material', 'descricao',
    'tipo_montagem', 'cap_montagem', 'cap_palmilha', 'cap_sola', 'cap_pintura',
    't_horizontal', 't_tiras', 't_vertical', 't_rotativa', 't_eva',
]
_COLS_OBRIG  = ['ref_linha', 'referencia', 'linha', 'material', 'descricao',
                'tipo_montagem', 'cap_montagem']

# Faixas válidas para valores de referência (min, max, unidade)
# cap_montagem: capacidade diária por referência/linha (pares/dia), não teto semanal
_FAIXAS_REFS = {
    'cap_montagem':  (100,   15000, 'pares/dia'),
    't_horizontal':  (0.50,  3.00,  'min/par'),
    't_tiras':       (0.10,  1.00,  'min/par'),
    't_vertical':    (0.10,  2.00,  'min/par'),
    't_rotativa':    (0.20,  4.00,  'min/par'),
    't_eva':         (0.20,  3.00,  'min/par'),
}
_FAIXAS_EFIC = (0.40, 1.00)
_FAIXAS_HORA = (4.0, 24.0)


def _xls_celula(ws, row, col, valor, bold=False, cor_fundo=None, cor_fonte='000000',
                align='left', tamanho=10, wrap=False):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        pass
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, color=cor_fonte, name='Calibri', size=tamanho)
    if cor_fundo:
        c.fill = PatternFill('solid', fgColor=cor_fundo)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c


def _xls_linha_header(ws, row, valores, cor_fundo='1C2030', cor_fonte='F3F0EB'):
    for col, v in enumerate(valores, 1):
        _xls_celula(ws, row, col, v, bold=True, cor_fundo=cor_fundo, cor_fonte=cor_fonte,
                    align='center')


def _sheet_leia_me(wb):
    ws = wb.create_sheet('LEIA-ME')
    ws.column_dimensions['A'].width = 120
    linhas = [
        ('BOAONDA INTELLIGENCE — Capacidade Fabril', True),
        ('Arquivo gerado automaticamente. Edite com cuidado.', False),
        ('', False),
        ('COMO USAR', True),
        ('1. Edite a aba PARÂMETROS para ajustar tetos do atelier, pools de injetados e eficiências.', False),
        ('2. Edite a aba REFERÊNCIAS para adicionar, remover ou ajustar referências.', False),
        ('3. Salve o arquivo e importe pelo portal (botão "Importar Excel" na tela de Capacidade Fabril).', False),
        ('4. O portal validará o arquivo antes de salvar — erros são exibidos com detalhes.', False),
        ('', False),
        ('ABA PARÂMETROS — CHAVES EDITÁVEIS', True),
        ('  atelier.convencional_pares_semana  — teto semanal do atelier convencional (pares)', False),
        ('  atelier.montado_pares_semana       — teto semanal do atelier montado (pares)', False),
        ('  pool.<nome>.maquinas               — número de máquinas ativas no pool', False),
        ('  pool.<nome>.horas_dia              — horas de operação por dia (validado: 4,0–24,0)', False),
        ('  pool.<nome>.min_dia_total          — capacidade total min/dia (editável diretamente)', False),
        ('  eficiencia.<nome>                  — eficiência efetiva (validado: 0,40–1,00)', False),
        ('  setor.filial_palmilha.capacidade_dia — capacidade diária Filial Palmilha (pares/dia)', False),
        ('  setor.filial_sola.capacidade_dia   — capacidade diária Filial Sola (pares/dia)', False),
        ('', False),
        ('ABA REFERÊNCIAS — COLUNAS', True),
        ('  ref_linha*    — chave única da linha (ex: "1913 KIN/102")  OBRIGATÓRIO', False),
        ('  referencia*   — código da referência (ex: "1913 KIN")      OBRIGATÓRIO', False),
        ('  linha*        — número da linha de produção                 OBRIGATÓRIO', False),
        ('  material*     — material (ex: TR, EVA, PU)                  OBRIGATÓRIO', False),
        ('  descricao*    — descrição do produto                        OBRIGATÓRIO', False),
        ('  tipo_montagem* — CONVENCIONAL ou MONTADO                   OBRIGATÓRIO', False),
        ('  cap_montagem* — capacidade de montagem (pares/dia, 100–15.000) OBRIGATÓRIO', False),
        ('  cap_palmilha  — capacidade Filial Palmilha (pares/dia) — 0 se não usa', False),
        ('  cap_sola      — capacidade Filial Sola (pares/dia) — 0 se não usa', False),
        ('  cap_pintura   — capacidade de pintura (pares/dia) — 0 se não usa', False),
        ('  t_horizontal  — tempo padrão pool horizontal (min/par) — 0 se não usa', False),
        ('  t_tiras       — tempo padrão pool tiras (min/par) — 0 se não usa', False),
        ('  t_vertical    — tempo padrão pool vertical (min/par) — 0 se não usa', False),
        ('  t_rotativa    — tempo padrão pool rotativa (min/par) — 0 se não usa', False),
        ('  t_eva         — tempo padrão pool EVA (min/par) — 0 se não usa', False),
        ('', False),
        ('REGRAS DE VALIDAÇÃO', True),
        ('  cap_montagem: 100 a 15.000 pares/dia', False),
        ('  t_horizontal: 0,50 a 3,00 min/par (quando preenchido > 0)', False),
        ('  t_tiras: 0,10 a 1,00 min/par (quando preenchido > 0)', False),
        ('  t_vertical: 0,10 a 2,00 min/par (quando preenchido > 0)', False),
        ('  t_rotativa: 0,20 a 4,00 min/par (quando preenchido > 0)', False),
        ('  t_eva: 0,20 a 3,00 min/par (quando preenchido > 0)', False),
        ('  horas/dia (pools): 4,0 a 24,0', False),
        ('  eficiências: 0,40 a 1,00', False),
    ]
    for i, (texto, bold) in enumerate(linhas, 1):
        _xls_celula(ws, i, 1, texto, bold=bold,
                    cor_fundo='1C2030' if bold else None,
                    cor_fonte='F3F0EB' if bold else '000000',
                    tamanho=11 if bold else 10, wrap=True)


def _sheet_parametros(wb, d):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet('PARÂMETROS')
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 22

    _xls_linha_header(ws, 1, ['CHAVE', 'VALOR', 'DESCRIÇÃO'])

    pools_cfg = d.get('pool_maquinas_injetados', {})
    efic_cfg  = d.get('eficiencias_padrao', {})
    atelier   = d.get('teto_atelier_semanal', {})
    setores   = d.get('setores_apoio', {})

    linhas = []
    # Atelier
    linhas.append(('atelier.convencional_pares_semana',
                   atelier.get('convencional_pares_semana', 29000),
                   'Teto convencional (pares/semana)'))
    linhas.append(('atelier.montado_pares_semana',
                   atelier.get('montado_pares_semana', 6000),
                   'Teto montado (pares/semana)'))
    # Pools
    for pool in _POOLS_ORDEM:
        cfg = pools_cfg.get(pool, {})
        linhas.append((f'pool.{pool}.maquinas',
                       cfg.get('maquinas', 0), f'Máquinas ativas — {pool}'))
        linhas.append((f'pool.{pool}.horas_dia',
                       cfg.get('horas_dia', 0), f'Horas/dia — {pool} (4,0–24,0)'))
        linhas.append((f'pool.{pool}.min_dia_total',
                       cfg.get('min_dia_total', 0), f'Capacidade total min/dia — {pool}'))
    # Eficiências
    for k in _EFIC_ORDEM:
        v_efic = efic_cfg.get(k)
        if v_efic is not None:
            linhas.append((f'eficiencia.{k}', v_efic, f'Eficiência {k} (0,40–1,00)'))
    # Setores
    palmilha = setores.get('filial_palmilha', {})
    sola     = setores.get('filial_sola', {})
    linhas.append(('setor.filial_palmilha.capacidade_dia',
                   palmilha.get('capacidade_dia', 0), 'Capacidade Filial Palmilha (pares/dia)'))
    linhas.append(('setor.filial_sola.capacidade_dia',
                   sola.get('capacidade_dia', 0), 'Capacidade Filial Sola (pares/dia)'))

    for i, (chave, valor, desc) in enumerate(linhas, 2):
        cor = 'F5F5F5' if i % 2 == 0 else None
        _xls_celula(ws, i, 1, chave, cor_fundo=cor)
        _xls_celula(ws, i, 2, valor, cor_fundo=cor, align='center')
        _xls_celula(ws, i, 3, desc, cor_fundo=cor)

    ws.freeze_panes = 'A2'


def _sheet_referencias(wb, d):
    ws = wb.create_sheet('REFERÊNCIAS')
    for col, nome in enumerate(_COLS_REFS, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 16
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['E'].width = 28

    _xls_linha_header(ws, 1, _COLS_REFS)

    refs = d.get('referencias', {})
    for i, (ref_linha, v) in enumerate(refs.items(), 2):
        cor = 'F9F9F9' if i % 2 == 0 else None
        # Flatten pool tempos
        tempos = v.get('tempos_padrao_pool_min', {})

        def t(pool):
            comp = tempos.get(pool, [])
            return round(sum(c['tempo_padrao_min'] for c in comp), 2) if comp else 0

        cap_apoio = v.get('capacidades_apoio', {})
        row = [
            ref_linha,
            v.get('referencia', ''),
            v.get('linha', ''),
            v.get('material', ''),
            v.get('descricao', ''),
            v.get('tipo_montagem', ''),
            v.get('capacidade_montagem_dia', ''),
            cap_apoio.get('palmilha') or 0,
            cap_apoio.get('sola') or 0,
            cap_apoio.get('pintura') or 0,
            t('horizontal'), t('tiras'), t('vertical'), t('rotativa'), t('eva'),
        ]
        for col, val in enumerate(row, 1):
            _xls_celula(ws, i, col, val, cor_fundo=cor)

    ws.freeze_panes = 'A2'


def exportar_capacidade_excel(dados_path):
    """Lê dados_capacidade.json e retorna BytesIO com o Excel gerado."""
    from io import BytesIO
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError('openpyxl não instalado. Rode: pip install openpyxl')

    with open(dados_path, encoding='utf-8') as f:
        d = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_leia_me(wb)
    _sheet_parametros(wb, d)
    _sheet_referencias(wb, d)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _ler_parametros(ws_params):
    """Lê a aba PARÂMETROS como dict chave→valor."""
    params = {}
    for row in ws_params.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            params[str(row[0]).strip()] = row[1]
    return params


def _validar_formato(wb):
    """Camada 1: verifica sheets e colunas obrigatórias."""
    erros = []
    for sheet in ('PARÂMETROS', 'REFERÊNCIAS'):
        if sheet not in wb.sheetnames:
            erros.append(f'Aba "{sheet}" não encontrada.')
    if erros:
        return erros

    ws = wb['REFERÊNCIAS']
    header = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(max_row=1))]
    for col in _COLS_OBRIG:
        if col.lower() not in header:
            erros.append(f'Coluna obrigatória ausente em REFERÊNCIAS: {col}')
    return erros


def _validar_faixas(wb):
    """Camada 2: valida faixas de valores em PARÂMETROS e REFERÊNCIAS."""
    erros = []
    params = _ler_parametros(wb['PARÂMETROS'])

    # Validar horas_dia e eficiências
    for pool in _POOLS_ORDEM:
        chave = f'pool.{pool}.horas_dia'
        v = params.get(chave)
        if v is not None:
            try:
                horas = float(v)
                lo, hi = _FAIXAS_HORA
                if not (lo <= horas <= hi):
                    erros.append(f'PARÂMETROS › {chave} = {horas}: esperado {lo}–{hi} h/dia')
            except (TypeError, ValueError):
                erros.append(f'PARÂMETROS › {chave}: valor inválido "{v}"')

    for k in _EFIC_ORDEM:
        chave = f'eficiencia.{k}'
        v = params.get(chave)
        if v is not None:
            try:
                efic = float(v)
                lo, hi = _FAIXAS_EFIC
                if not (lo <= efic <= hi):
                    erros.append(f'PARÂMETROS › {chave} = {efic}: esperado {lo}–{hi}')
            except (TypeError, ValueError):
                erros.append(f'PARÂMETROS › {chave}: valor inválido "{v}"')

    # Validar cada referência
    ws = wb['REFERÊNCIAS']
    header = [str(c.value or '').strip().lower()
              for c in next(ws.iter_rows(max_row=1))]
    col_idx = {nome: i for i, nome in enumerate(header)}

    for n_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or row[0] is None:
            continue
        ref_linha = str(row[0]).strip()
        for campo, (lo, hi, _unidade) in _FAIXAS_REFS.items():
            idx = col_idx.get(campo)
            if idx is None:
                continue
            raw = row[idx]
            if raw is None or raw == '' or raw == 0:
                continue  # campo vazio/zero = não usa pool, OK
            try:
                val = float(raw)
            except (TypeError, ValueError):
                erros.append(f'REFERÊNCIAS › linha {n_row} ({ref_linha}) › {campo}: '
                             f'valor não numérico "{raw}"')
                continue
            if not (lo <= val <= hi):
                erros.append(f'REFERÊNCIAS › linha {n_row} ({ref_linha}) › {campo} = {val}: '
                             f'esperado {lo}–{hi}')
    return erros


def _reconstruir_json(wb, dados_atual, output_dir):
    """Reconstrói dados_capacidade.json a partir do Excel importado."""
    params = _ler_parametros(wb['PARÂMETROS'])

    def p(chave, default=None):
        v = params.get(chave)
        return float(v) if v is not None else default

    def pi(chave, default=0):
        v = params.get(chave)
        return int(float(v)) if v is not None else default

    # Teto atelier
    conv = pi('atelier.convencional_pares_semana', 29000)
    mont = pi('atelier.montado_pares_semana', 6000)
    teto_atelier = {
        'convencional_pares_semana': conv,
        'montado_pares_semana': mont,
        'total_pares_semana': conv + mont,
        'observacao': dados_atual.get('teto_atelier_semanal', {}).get('observacao', ''),
    }

    # Pools
    pools = {}
    for pool in _POOLS_ORDEM:
        cfg_atual = dados_atual.get('pool_maquinas_injetados', {}).get(pool, {})
        maquinas  = pi(f'pool.{pool}.maquinas', cfg_atual.get('maquinas', 0))
        horas_dia = p(f'pool.{pool}.horas_dia', cfg_atual.get('horas_dia', 0))
        min_total = p(f'pool.{pool}.min_dia_total', cfg_atual.get('min_dia_total',
                                                                    round(maquinas * horas_dia * 60, 1)))
        entry = {'maquinas': maquinas, 'horas_dia': horas_dia, 'min_dia_total': min_total}
        # Preserva campos extras (maquinas_ativas, maquina_reserva)
        for k, v in cfg_atual.items():
            if k not in entry:
                entry[k] = v
        pools[pool] = entry

    # Eficiências
    efic = {}
    for k in _EFIC_ORDEM:
        v = p(f'eficiencia.{k}')
        if v is not None:
            efic[k] = v
    # Preserva campos extras do JSON atual (observações, etc.)
    for k, v in dados_atual.get('eficiencias_padrao', {}).items():
        if k not in efic:
            efic[k] = v

    # Referências
    ws = wb['REFERÊNCIAS']
    header = [str(c.value or '').strip().lower()
              for c in next(ws.iter_rows(max_row=1))]
    col_idx = {nome: i for i, nome in enumerate(header)}

    def cv(row, campo, default=None):
        idx = col_idx.get(campo)
        if idx is None:
            return default
        v = row[idx]
        return v if v is not None and v != '' else default

    def cv_float(row, campo, default=None):
        v = cv(row, campo)
        try:
            return float(v) if v is not None else default
        except (TypeError, ValueError):
            return default

    def cv_int(row, campo, default=None):
        v = cv_float(row, campo)
        return int(v) if v is not None else default

    referencias = {}
    por_tipo = Counter()
    palmilha_refs = sola_refs = 0
    palmilha_caps: set = set()
    sola_caps: set = set()
    pintura_grupos = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        ref_linha = str(row[0]).strip()
        if not ref_linha:
            continue

        tipo_raw = str(cv(row, 'tipo_montagem') or '').strip().upper()
        tipo = 'MONTADO' if 'MONTADO' in tipo_raw else 'CONVENCIONAL'
        por_tipo[tipo] += 1

        cap_mont = cv_int(row, 'cap_montagem')
        cap_palmilha = cv_int(row, 'cap_palmilha')
        cap_sola     = cv_int(row, 'cap_sola')
        cap_pintura  = cv_int(row, 'cap_pintura')

        if cap_palmilha:
            palmilha_refs += 1
            palmilha_caps.add(cap_palmilha)
        if cap_sola:
            sola_refs += 1
            sola_caps.add(cap_sola)
        if cap_pintura:
            pintura_grupos[cap_pintura] += 1

        tempos = {}
        for pool, campo in [('horizontal', 't_horizontal'), ('tiras', 't_tiras'),
                             ('vertical', 't_vertical'), ('rotativa', 't_rotativa'),
                             ('eva', 't_eva')]:
            t = cv_float(row, campo)
            if t and t > 0:
                tempos[pool] = [{'componente': pool, 'tempo_padrao_min': round(t, 2)}]

        referencias[ref_linha] = {
            'material':   str(cv(row, 'material') or '').strip(),
            'referencia': str(cv(row, 'referencia') or '').strip(),
            'linha':      cv(row, 'linha'),
            'descricao':  str(cv(row, 'descricao') or '').strip(),
            'tipo_montagem': tipo,
            'capacidade_montagem_dia': cap_mont,
            'capacidades_apoio': {
                'palmilha': cap_palmilha or None,
                'sola':     cap_sola or None,
                'pintura':  cap_pintura or None,
            },
            'tempos_padrao_pool_min': tempos,
        }

    # Setores de apoio
    setores_atual = dados_atual.get('setores_apoio', {})
    cap_palh = pi('setor.filial_palmilha.capacidade_dia',
                  setores_atual.get('filial_palmilha', {}).get('capacidade_dia', 0))
    cap_sola_val = pi('setor.filial_sola.capacidade_dia',
                      setores_atual.get('filial_sola', {}).get('capacidade_dia', 0))

    setores = {
        'filial_palmilha': {
            'capacidade_dia': max(palmilha_caps) if palmilha_caps else cap_palh,
            'refs_que_usam': palmilha_refs,
            'observacao': setores_atual.get('filial_palmilha', {}).get('observacao', ''),
        },
        'filial_sola': {
            'capacidade_dia': max(sola_caps) if sola_caps else cap_sola_val,
            'refs_que_usam': sola_refs,
            'observacao': setores_atual.get('filial_sola', {}).get('observacao', ''),
        },
        'pintura': {
            'grupos': {str(k): v for k, v in sorted(pintura_grupos.items())},
            'total_refs': sum(pintura_grupos.values()),
            'observacao': setores_atual.get('pintura', {}).get('observacao', ''),
        },
        'giro_dias': setores_atual.get('giro_dias', GIRO_DIAS),
        'capacidade_dia_observada': setores_atual.get('capacidade_dia_observada',
                                                       CAPACIDADE_DIA_OBSERVADA),
    }

    metadados = {
        'total_referencias_cadastradas': len(referencias),
        'total_com_capacidade_montagem': len(referencias),
        'por_tipo': dict(por_tipo),
    }

    validacao_cobertura = calcular_validacao_cobertura(referencias, output_dir)

    return {
        'gerado_em': datetime.now().strftime('%d/%m/%Y'),
        'teto_atelier_semanal': teto_atelier,
        'pool_maquinas_injetados': pools,
        'eficiencias_padrao': efic,
        'metadados': metadados,
        'referencias': referencias,
        'setores_apoio': setores,
        'validacao_cobertura': validacao_cobertura,
    }


def importar_capacidade_excel(path_excel, dados_path, output_dir=None):
    """Importa Excel com 3 camadas de proteção e sobrescreve dados_capacidade.json.

    Retorna dict com 'status' (ok/erro) e detalhes.
    """
    try:
        import openpyxl
    except ImportError:
        return {'status': 'erro', 'mensagem': 'openpyxl não instalado.'}

    if output_dir is None:
        output_dir = os.path.dirname(dados_path)

    wb = openpyxl.load_workbook(path_excel, data_only=True)

    # ── Camada 1: formato ─────────────────────────────────────────────
    erros_fmt = _validar_formato(wb)
    if erros_fmt:
        return {'status': 'erro', 'mensagem': 'Falha na validação de formato:\n' + '\n'.join(erros_fmt)}

    # ── Camada 2: faixas de valores ───────────────────────────────────
    erros_faixas = _validar_faixas(wb)
    if erros_faixas:
        return {'status': 'erro',
                'mensagem': 'Falha na validação de faixas:\n' + '\n'.join(erros_faixas)}

    # ── Camada 3: backup + reconstrução + gravação ────────────────────
    dados_atual = {}
    if os.path.exists(dados_path):
        with open(dados_path, encoding='utf-8') as f:
            dados_atual = json.load(f)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(output_dir,
                                   f'dados_capacidade_backup_{ts}.json')
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump(dados_atual, f, ensure_ascii=False, default=str)

    novo = _reconstruir_json(wb, dados_atual, output_dir)

    with open(dados_path, 'w', encoding='utf-8') as f:
        json.dump(novo, f, ensure_ascii=False, default=str)

    return {
        'status': 'ok',
        'refs_importadas': len(novo['referencias']),
        'por_tipo': novo['metadados']['por_tipo'],
        'cobertura_pct': novo['validacao_cobertura'].get('pct_cobertura_volume'),
        'gerado_em': novo['gerado_em'],
    }


# ─── CLI ─────────────────────────────────────────────────────────────
def main():
    print("=" * 52)
    print("  BOAONDA INTELLIGENCE — Capacidade Fabril")
    print("=" * 52)

    caminho = achar_planilha_capacidade('.')
    if not caminho:
        print("\n  ✗ Planilha 'Programação_GERAL' (.xlsx) não encontrada na pasta atual.")
        print("    Coloque o arquivo aqui e rode novamente.\n")
        sys.exit(1)

    print(f"\n  ✓ {caminho} ({os.path.getsize(caminho)/1024:.0f} KB)")
    dados = processar_capacidade(caminho, output_dir='.')

    v = dados['validacao_cobertura']
    print("\n" + "=" * 52)
    print("  RESUMO")
    print("=" * 52)
    print(f"\n  {dados['metadados']['total_com_capacidade_montagem']} referências mapeadas "
          f"({dados['metadados']['por_tipo']})")
    if v.get('pct_cobertura_volume') is not None:
        print(f"  Cobertura da programação real: {v['pct_cobertura_volume']}% "
              f"({v['refs_com_capacidade']}/{v['refs_programacao_unicas']} refs)")
        if v['refs_sem_capacidade']:
            print(f"  Refs sem capacidade: {', '.join(r['referencia'] for r in v['refs_sem_capacidade'])}")
    print(f"\n  Recarregue o portal no browser (F5).")
    print("=" * 52 + "\n")


if __name__ == '__main__':
    main()
