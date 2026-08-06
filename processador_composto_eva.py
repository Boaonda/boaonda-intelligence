"""
Boaonda Intelligence — Composto EVA: demanda interna
=====================================================
O setor de Composto de EVA atende dois blocos:

  1. VENDA EXTERNA — composto faturado direto a terceiros (Ramarim, M3, Kidy...).
     Já coberto por `processar_vendas_eva()` em processador.py e pelo módulo
     de Faturamento. Identificado por marca == 'COMPOSTOS EVA'.

  2. DEMANDA INTERNA (este módulo) — o composto consumido pela própria
     Boaonda, porque todo calçado injetado em EVA gasta composto produzido
     por este setor. Até 2026-08 não existia nenhum controle disso: o setor
     não enxergava quanto precisava produzir para a fábrica, nem com quanta
     antecedência.

Como a demanda interna é calculada
----------------------------------
A programação de produção está em PARES; o composto trabalha em KG. A ponte
é um fator kg/par por referência, que **não existe em nenhuma base** — o 3YS
não tem coluna de peso (88 colunas conferidas) e o cadastro de Capacidade
Fabril também não. Enquanto o T.I. não disponibiliza a ficha técnica/BOM do
ERP, o fator é mantido à mão via round-trip de Excel (mesmo padrão de
Capacidade Fabril e Metas Comerciais): `dados_composto_fatores.json`.

"É EVA?" vem do MATERIAL do cadastro técnico
--------------------------------------------
Existem duas definições de EVA no portal, e elas divergem:

  - `Linha` (3YS)  = linha COMERCIAL (CLASSIC/EVA/WORKS/FIT/DAY BY DAY) —
    é a do gráfico "Mix de linhas" e da regra de lote mínimo.
  - `material` (cadastro de Capacidade, por referência+linha) = material de
    INJEÇÃO (EVA/TR) — o que de fato determina consumo de composto.

Medido na janela real da programação (353 mil pares): 11.577 pares são
material EVA mas estão na linha comercial CLASSIC (ex.: 2602 IVY). O
contrário nunca ocorre — ou seja, usar a linha comercial SUBESTIMA o
consumo. Decisão do Cássio em 2026-08-06: usar sempre o **material** do
cadastro técnico. Ver [[project_composto_eva_demanda_interna]].

Saída: dados_composto_demanda.json
"""

import json
import os
from collections import defaultdict
from datetime import datetime, timedelta

FATORES_JSON = 'dados_composto_fatores.json'
DEMANDA_JSON = 'dados_composto_demanda.json'

# Faixa aceita para o fator kg/par na importação. Um par de sandália de EVA
# fica na casa das centenas de gramas; a faixa é larga de propósito (só barra
# erro grosseiro de digitação, tipo gramas no lugar de quilos).
KG_PAR_MIN = 0.01
KG_PAR_MAX = 20.0


# ─── HELPERS ─────────────────────────────────────────────────────────
def _num(v):
    """Converte célula/valor em float, aceitando vírgula decimal. None se vazio."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '.')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def cor_injecao_de(cor):
    """Cor de injeção = primeira cor antes da '/' (mesma regra do quadro de
    lote mínimo em Programação: 'CANELA/ROSE NUDE' → injeta CANELA)."""
    return ((cor or '').split('/')[0] or '').strip().upper() or '(sem cor)'


def mes_ref_da_semana(semana_key):
    """Mês de referência de uma semana = mês da SEXTA-FEIRA daquela semana.
    Mesma regra do mesRef() em boaonda_programacao_v3.html, para que os
    totais mensais daqui batam com os da tela de Programação."""
    try:
        ini = datetime.strptime(semana_key, '%Y-%m-%d')
    except ValueError:
        return semana_key[:7]
    sexta = ini + timedelta(days=4)
    return sexta.strftime('%Y-%m')


def carregar_material_map(capacidade_path):
    """Lê dados_capacidade.json e devolve (por_ref_linha, por_ref).

    por_ref_linha: {(REF, LINHA): 'EVA'|'TR'|...}  — chave exata
    por_ref:       {REF: material predominante}    — fallback quando a
                   combinação (referência, linha) não está cadastrada
                   (medido: ~35 mil pares da janela caem nesse caso).
    """
    por_ref_linha = {}
    contagem = defaultdict(lambda: defaultdict(int))
    if not capacidade_path or not os.path.exists(capacidade_path):
        return por_ref_linha, {}
    with open(capacidade_path, encoding='utf-8') as f:
        cap = json.load(f)
    for v in (cap.get('referencias') or {}).values():
        ref = str(v.get('referencia', '')).strip().upper()
        linha = str(v.get('linha', '')).strip()
        material = str(v.get('material') or '').strip().upper()
        if not ref or not material:
            continue
        por_ref_linha[(ref, linha)] = material
        contagem[ref][material] += 1
    por_ref = {ref: max(m.items(), key=lambda x: x[1])[0] for ref, m in contagem.items()}
    return por_ref_linha, por_ref


def material_de(por_ref_linha, por_ref, ref, linha):
    """Material de injeção de um item da programação, com fallback por
    referência quando a linha específica não está cadastrada."""
    ref = (ref or '').strip().upper()
    linha = str(linha or '').strip()
    m = por_ref_linha.get((ref, linha))
    if m:
        return m
    return por_ref.get(ref)


def carregar_fatores(fatores_path):
    """Lê dados_composto_fatores.json → {REF: {'': kg_par, '110': kg_par}}.

    A chave '' é o fator que vale para TODAS as linhas daquela referência;
    uma chave de linha específica sobrescreve. Isso mantém a planilha simples
    hoje (um número por referência) sem impedir precisão por linha depois,
    quando a ficha técnica do ERP chegar.
    """
    if not fatores_path or not os.path.exists(fatores_path):
        return {}
    with open(fatores_path, encoding='utf-8') as f:
        d = json.load(f)
    out = defaultdict(dict)
    for item in (d.get('fatores') or []):
        ref = str(item.get('referencia', '')).strip().upper()
        linha = str(item.get('linha') or '').strip()
        kg = _num(item.get('kg_par'))
        if ref and kg and kg > 0:
            out[ref][linha] = kg
    return dict(out)


def resolver_fator(fatores, ref, linha):
    """Fator kg/par de um item: linha específica primeiro, senão o da
    referência inteira. Devolve (kg_par, origem) — origem em ('linha',
    'referencia', None) para o relatório de cobertura ser honesto sobre de
    onde veio cada número."""
    ref = (ref or '').strip().upper()
    linha = str(linha or '').strip()
    porref = fatores.get(ref)
    if not porref:
        return None, None
    if linha and linha in porref:
        return porref[linha], 'linha'
    if '' in porref:
        return porref[''], 'referencia'
    return None, None


# ─── DEMANDA INTERNA ─────────────────────────────────────────────────
def processar_demanda_composto(detalhe_path, capacidade_path, fatores_path, output_dir='.'):
    """Converte a programação de calçado EVA (pares) em necessidade de
    composto (kg), por semana e por mês, com quebra por cor de injeção.

    Cobre o histórico INTEIRO da programação (via dados_composto_base.json),
    diferente dos drilldowns da tela de Programação, que ficam limitados à
    janela rolante de ~2 meses do detalhe item-a-item.
    """
    print("\n  Processando demanda interna de Composto EVA...")

    # Fonte preferencial: dados_composto_base.json — histórico INTEIRO da
    # programação, compacto (só ref/linha/cor/linha_prod por semana). O
    # detalhe item-a-item serve de fallback, mas é uma janela rolante de
    # ~2 meses: usá-lo cortaria a visão do período, que é justamente o que
    # o setor precisa enxergar por completo.
    base_path = os.path.join(os.path.dirname(detalhe_path) or '.', 'dados_composto_base.json')
    origem_dados = None
    if os.path.exists(base_path):
        with open(base_path, encoding='utf-8') as f:
            det = json.load(f)
        origem_dados = 'base'
    elif os.path.exists(detalhe_path):
        with open(detalhe_path, encoding='utf-8') as f:
            det = json.load(f)
        origem_dados = 'detalhe'
        print("    ⚠ dados_composto_base.json ausente — usando o detalhe da programação "
              "(janela de ~2 meses). Reprocesse os dados para cobrir o histórico completo.")
    else:
        print("    ⚠ Nem dados_composto_base.json nem dados_programacao_detalhe.json "
              "encontrados — pulando.")
        return None

    por_ref_linha, por_ref = carregar_material_map(capacidade_path)
    fatores = carregar_fatores(fatores_path)

    if not por_ref_linha:
        print("    ⚠ dados_capacidade.json ausente — sem classificação de material, "
              "não dá para identificar o que é EVA.")

    # semanas[wk] = {kg, pares, pares_sem_fator, cores{}, refs{}}
    def _novo_balde():
        return {'kg': 0.0, 'pares': 0, 'pares_sem_fator': 0,
                'cores': defaultdict(lambda: {'kg': 0.0, 'pares': 0,
                                              'refs': defaultdict(lambda: {'kg': 0.0, 'pares': 0})}),
                'refs': defaultdict(lambda: {'kg': 0.0, 'pares': 0})}

    semanas = defaultdict(_novo_balde)
    meses = defaultdict(_novo_balde)

    pares_eva_total = 0
    pares_com_fator = 0
    refs_sem_fator = defaultdict(int)
    origem_fator = defaultdict(int)
    # Referência que a programação tem mas o cadastro técnico não conhece:
    # sem material, não dá para saber se é EVA. Antes isso era descartado em
    # silêncio — é justamente onde um PRODUTO NOVO DE EVA cairia, sumindo da
    # necessidade sem nenhum aviso. Agora é reportado, e quando a linha
    # comercial diz EVA vira alerta explícito na tela.
    refs_desconhecidas = defaultdict(lambda: {'pares': 0, 'linha_eva': False})

    for wk, itens in (det.get('semanas') or {}).items():
        mk = mes_ref_da_semana(wk)
        for it in itens:
            ref = (it.get('ref') or '').strip().upper()
            linha = str(it.get('linha') or '').strip()
            pares = it.get('pares') or 0
            if pares <= 0:
                continue
            material = material_de(por_ref_linha, por_ref, ref, linha)
            if material is None:
                d = refs_desconhecidas[ref]
                d['pares'] += pares
                if (it.get('linha_prod') or '').strip().upper() == 'EVA':
                    d['linha_eva'] = True
                continue
            if material != 'EVA':
                continue

            pares_eva_total += pares
            kg_par, origem = resolver_fator(fatores, ref, linha)
            if kg_par:
                pares_com_fator += pares
                origem_fator[origem] += pares
                kg = pares * kg_par
            else:
                refs_sem_fator[ref] += pares
                kg = 0.0

            cor = cor_injecao_de(it.get('cor'))
            for balde in (semanas[wk], meses[mk]):
                balde['kg'] += kg
                balde['pares'] += pares
                if not kg_par:
                    balde['pares_sem_fator'] += pares
                c = balde['cores'][cor]
                c['kg'] += kg
                c['pares'] += pares
                cr = c['refs'][ref]
                cr['kg'] += kg
                cr['pares'] += pares
                r = balde['refs'][ref]
                r['kg'] += kg
                r['pares'] += pares

    def _serializa(balde):
        cores = sorted(
            [{'cor': cor, 'kg': round(v['kg'], 1), 'pares': v['pares'],
              'refs': sorted(
                  [{'ref': r, 'kg': round(rv['kg'], 1), 'pares': rv['pares']}
                   for r, rv in v['refs'].items()],
                  key=lambda x: -x['pares'])}
             for cor, v in balde['cores'].items()],
            key=lambda x: -x['pares'])
        refs = sorted(
            [{'ref': r, 'kg': round(v['kg'], 1), 'pares': v['pares']}
             for r, v in balde['refs'].items()],
            key=lambda x: -x['pares'])
        return {'kg': round(balde['kg'], 1), 'pares': balde['pares'],
                'pares_sem_fator': balde['pares_sem_fator'],
                'cores': cores, 'refs': refs}

    out = {
        'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'cutoff': det.get('cutoff'),
        # 'base' = histórico completo | 'detalhe' = janela rolante de ~2 meses.
        # Não confundir com 'pares_por_origem_do_fator' abaixo, que fala de
        # onde veio o kg/par (linha específica ou referência inteira).
        'origem_dados': origem_dados,
        'semanas': {k: _serializa(v) for k, v in sorted(semanas.items())},
        'meses': {k: _serializa(v) for k, v in sorted(meses.items())},
        'cobertura': {
            'pares_eva': pares_eva_total,
            'pares_com_fator': pares_com_fator,
            'pct_com_fator': round(pares_com_fator / pares_eva_total * 100, 1) if pares_eva_total else 0.0,
            'pares_por_origem_do_fator': dict(origem_fator),
            'refs_sem_fator': sorted(
                [{'ref': r, 'pares': p} for r, p in refs_sem_fator.items()],
                key=lambda x: -x['pares']),
            'refs_com_fator_cadastrado': len(fatores),
            'refs_material_desconhecido': sorted(
                [{'ref': r, 'pares': v['pares'], 'linha_comercial_eva': v['linha_eva']}
                 for r, v in refs_desconhecidas.items()],
                key=lambda x: -x['pares']),
        },
    }

    destino = os.path.join(output_dir, DEMANDA_JSON)
    with open(destino, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, default=str)

    kg_total = sum(v['kg'] for v in out['semanas'].values())
    print(f"    {pares_eva_total:,} pares EVA na janela | "
          f"{out['cobertura']['pct_com_fator']:.0f}% com fator kg/par | "
          f"{kg_total:,.0f} kg de necessidade")
    if refs_sem_fator:
        print(f"    ⚠ {len(refs_sem_fator)} referência(s) EVA sem fator cadastrado "
              f"({sum(refs_sem_fator.values()):,} pares não convertidos)")
    suspeitas = [r for r, v in refs_desconhecidas.items() if v['linha_eva']]
    if suspeitas:
        print(f"    ⚠ {len(suspeitas)} referência(s) fora do cadastro técnico com linha "
              f"comercial EVA — possível produto novo: {', '.join(sorted(suspeitas)[:5])}")
    print(f"    ✓ {DEMANDA_JSON} gerado")
    return out


# ─── EXCEL: EXPORTAR (round-trip, igual Capacidade/Metas) ────────────
def _xls_cel(ws, row, col, valor, bold=False, fundo=None, fonte='000000',
             align='left', tamanho=10, wrap=False, fmt=None):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, color=fonte, name='Calibri', size=tamanho)
    if fundo:
        c.fill = PatternFill('solid', fgColor=fundo)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    return c


def _sheet_leia_me_fatores(wb):
    ws = wb.create_sheet('LEIA-ME')
    ws.column_dimensions['A'].width = 118
    linhas = [
        ('BOAONDA INTELLIGENCE — Composto EVA: fator kg/par', True),
        ('', False),
        ('PARA QUE SERVE', True),
        ('A programação de produção está em PARES, mas o setor de Composto trabalha em KG.', False),
        ('Este arquivo guarda quantos quilos de composto cada par consome — é o que permite', False),
        ('traduzir a programação de calçado EVA em necessidade de composto.', False),
        ('', False),
        ('COMO PREENCHER', True),
        ('1. Preencha a coluna kg_par na aba FATORES (aceita vírgula ou ponto decimal).', False),
        ('2. As referências estão ordenadas por volume programado — as de cima pesam mais no total.', False),
        ('3. Salve o arquivo e importe pelo portal, na tela de Vendas Composto EVA.', False),
        ('4. Linhas deixadas em branco simplesmente não são convertidas; o portal mostra', False),
        ('   quanto do volume ficou sem fator, então dá para preencher aos poucos.', False),
        ('', False),
        ('COLUNAS DA ABA FATORES', True),
        ('  referencia*  — código da referência (ex.: "2402 EASY")            OBRIGATÓRIO', False),
        ('  linha        — número da linha/grade. DEIXE EM BRANCO para valer', False),
        ('                 para todas as linhas da referência (é o caso normal).', False),
        ('                 Preencha só se aquela linha específica tiver consumo diferente.', False),
        ('  kg_par       — quilos de composto por par (ex.: 0,25). Faixa aceita: '
         f'{KG_PAR_MIN} a {KG_PAR_MAX}.', False),
        ('  material / descricao / pares_programados — apenas informativos, não são lidos.', False),
        ('', False),
        ('OBSERVAÇÕES', True),
        ('  • "É EVA?" vem do MATERIAL do cadastro de Capacidade Fabril (EVA/TR), não da', False),
        ('    linha comercial (CLASSIC/EVA/WORKS/...). São coisas diferentes: há referências', False),
        ('    de linha comercial CLASSIC que são injetadas em EVA (ex.: 2602 IVY).', False),
        ('  • Por ora o consumo é considerado igual para todas as numerações da grade.', False),
        ('  • Quando a ficha técnica (BOM) do ERP ficar pronta, esta planilha pode ser', False),
        ('    substituída pela carga automática — o formato do JSON já suporta isso.', False),
    ]
    for i, (txt, bold) in enumerate(linhas, 1):
        _xls_cel(ws, i, 1, txt, bold=bold, tamanho=11 if bold else 10, wrap=False)
    return ws


def _levantar_refs_eva(capacidade_path, detalhe_path):
    """Monta a lista de referências EVA para a planilha: as do cadastro
    técnico (material == EVA) unidas às que aparecem na programação e
    resolvem como EVA — com o volume programado, para ordenar por relevância
    (quem preenche começa pelo que mais pesa)."""
    por_ref_linha, por_ref = carregar_material_map(capacidade_path)

    info = {}
    for (ref, linha), material in por_ref_linha.items():
        if material != 'EVA':
            continue
        d = info.setdefault(ref, {'referencia': ref, 'material': 'EVA',
                                  'linhas': set(), 'descricao': '', 'pares': 0})
        if linha:
            d['linhas'].add(linha)

    # descrições legíveis do cadastro
    if capacidade_path and os.path.exists(capacidade_path):
        with open(capacidade_path, encoding='utf-8') as f:
            cap = json.load(f)
        for v in (cap.get('referencias') or {}).values():
            ref = str(v.get('referencia', '')).strip().upper()
            if ref in info and not info[ref]['descricao']:
                info[ref]['descricao'] = str(v.get('descricao') or '')

    # volume programado (e refs que só aparecem na programação)
    if detalhe_path and os.path.exists(detalhe_path):
        with open(detalhe_path, encoding='utf-8') as f:
            det = json.load(f)
        for itens in (det.get('semanas') or {}).values():
            for it in itens:
                ref = (it.get('ref') or '').strip().upper()
                linha = str(it.get('linha') or '').strip()
                if material_de(por_ref_linha, por_ref, ref, linha) != 'EVA':
                    continue
                d = info.setdefault(ref, {'referencia': ref, 'material': 'EVA',
                                          'linhas': set(), 'descricao': '', 'pares': 0})
                d['pares'] += it.get('pares') or 0

    return sorted(info.values(), key=lambda x: (-x['pares'], x['referencia']))


def exportar_fatores_excel(capacidade_path, detalhe_path, fatores_path):
    """Gera o Excel de fatores kg/par já preenchido com o que existe hoje.
    Retorna BytesIO."""
    from io import BytesIO
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError('openpyxl não instalado. Rode: pip install openpyxl')

    refs = _levantar_refs_eva(capacidade_path, detalhe_path)
    fatores = carregar_fatores(fatores_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_leia_me_fatores(wb)

    ws = wb.create_sheet('FATORES')
    header = ['referencia*', 'linha', 'kg_par', 'material', 'descricao', 'pares_programados']
    for col, h in enumerate(header, 1):
        _xls_cel(ws, 1, col, h, bold=True, fundo='1C2030', fonte='F3F0EB', align='center')
    larguras = [26, 10, 12, 12, 42, 20]
    for col, w in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    linha_out = 2
    for r in refs:
        atual = fatores.get(r['referencia'], {})
        # Linha "curinga" da referência (linha em branco = vale para todas)
        _xls_cel(ws, linha_out, 1, r['referencia'])
        _xls_cel(ws, linha_out, 2, '', align='center')
        _xls_cel(ws, linha_out, 3, atual.get('', None), align='center',
                 fundo='FBE4D8', fmt='0.000')
        _xls_cel(ws, linha_out, 4, r['material'], align='center')
        _xls_cel(ws, linha_out, 5, r['descricao'])
        _xls_cel(ws, linha_out, 6, r['pares'], align='center', fmt='#,##0')
        linha_out += 1
        # Overrides por linha que já existirem no JSON são preservados
        for lin in sorted(k for k in atual if k):
            _xls_cel(ws, linha_out, 1, r['referencia'])
            _xls_cel(ws, linha_out, 2, lin, align='center')
            _xls_cel(ws, linha_out, 3, atual[lin], align='center', fundo='FBE4D8', fmt='0.000')
            _xls_cel(ws, linha_out, 4, r['material'], align='center')
            _xls_cel(ws, linha_out, 5, '(override desta linha)')
            _xls_cel(ws, linha_out, 6, '', align='center')
            linha_out += 1

    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── EXCEL: IMPORTAR ─────────────────────────────────────────────────
def _norm_header(v):
    return str(v or '').strip().lower().rstrip('*').strip()


def importar_fatores_excel(path_excel, fatores_path, output_dir=None):
    """Lê o Excel de fatores e sobrescreve dados_composto_fatores.json.
    Valida antes de gravar e faz backup do arquivo anterior."""
    try:
        import openpyxl
    except ImportError:
        return {'status': 'erro', 'mensagem': 'openpyxl não instalado.'}

    if output_dir is None:
        output_dir = os.path.dirname(fatores_path) or '.'

    wb = openpyxl.load_workbook(path_excel, data_only=True)
    if 'FATORES' not in wb.sheetnames:
        return {'status': 'erro',
                'mensagem': 'A aba "FATORES" não foi encontrada na planilha. '
                            'Use o arquivo baixado pelo próprio portal como base.'}
    ws = wb['FATORES']

    header = [_norm_header(c.value) for c in next(ws.iter_rows(max_row=1))]
    if 'referencia' not in header or 'kg_par' not in header:
        return {'status': 'erro',
                'mensagem': 'A aba FATORES precisa das colunas "referencia" e "kg_par".'}
    i_ref = header.index('referencia')
    i_kg = header.index('kg_par')
    i_lin = header.index('linha') if 'linha' in header else None

    fatores = []
    erros = []
    vistos = set()
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or i_ref >= len(row):
            continue
        ref = str(row[i_ref] or '').strip().upper()
        if not ref:
            continue
        linha = ''
        if i_lin is not None and i_lin < len(row):
            linha = str(row[i_lin] or '').strip()
            if linha.endswith('.0'):
                linha = linha[:-2]
        bruto = row[i_kg] if i_kg < len(row) else None
        kg = _num(bruto)
        if kg is None:
            continue  # em branco = não preenchido ainda, não é erro
        if kg < KG_PAR_MIN or kg > KG_PAR_MAX:
            erros.append(f'Linha {n} ({ref}): kg_par = {bruto} fora da faixa '
                         f'{KG_PAR_MIN}–{KG_PAR_MAX}.')
            continue
        chave = (ref, linha)
        if chave in vistos:
            erros.append(f'Linha {n}: referência "{ref}" '
                         f'{"linha " + linha if linha else "(todas as linhas)"} duplicada.')
            continue
        vistos.add(chave)
        fatores.append({'referencia': ref, 'linha': linha, 'kg_par': round(kg, 4)})

    if erros:
        return {'status': 'erro',
                'mensagem': 'A planilha tem erros e nada foi salvo:\n' + '\n'.join(erros[:20])}
    if not fatores:
        return {'status': 'erro',
                'mensagem': 'Nenhum fator kg/par preenchido na planilha — nada a importar.'}

    if os.path.exists(fatores_path):
        with open(fatores_path, encoding='utf-8') as f:
            anterior = f.read()
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        with open(os.path.join(output_dir, f'dados_composto_fatores_backup_{ts}.json'),
                  'w', encoding='utf-8') as f:
            f.write(anterior)

    saida = {
        'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'origem': 'planilha',
        'fatores': sorted(fatores, key=lambda x: (x['referencia'], x['linha'])),
    }
    with open(fatores_path, 'w', encoding='utf-8') as f:
        json.dump(saida, f, ensure_ascii=False, default=str)

    n_refs = len({x['referencia'] for x in fatores})
    return {'status': 'ok', 'fatores_importados': len(fatores), 'referencias': n_refs,
            'gerado_em': saida['gerado_em']}
